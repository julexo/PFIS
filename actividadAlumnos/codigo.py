import unittest
import pandas as pd
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, landscape
from reportlab.pdfgen import canvas
import os

# Función para extraer encabezados del archivo Excel
def extraer_encabezados(hoja):
    return {
        "asignatura": hoja["A1"].value,
        "fecha": hoja["A2"].value,
        "semana_docencia": hoja["A3"].value,
        "tema": hoja["A4"].value
    }

# Función para procesar datos del Excel
def procesar_datos(excel_file):
    pd.set_option('display.max_rows', None)  # Mostrar todas las filas
    df = pd.read_excel(excel_file, skiprows=4)  # Cambiado skiprows=5 a skiprows=4 para evitar perder datos
    df.columns = ["Nombre", "Apellido", "", "Nombre", "Apellido"]
    df = df.drop(columns=df.columns[2])
    df = df.fillna("")
    print("DataFrame cargado:")  # Depuración
    print(df)  # Esto imprimirá todas las filas sin truncar
    return df

# Función para generar PDF
def generar_pdf(df, encabezados, pdf_file):
    c = canvas.Canvas(pdf_file, pagesize=landscape(letter))
    width, height = landscape(letter)

    # Dibujar encabezado
    c.setFont("Helvetica-Bold", 14)
    y_offset = height - 40
    for key, value in encabezados.items():
        c.drawString(30, y_offset, str(value))
        y_offset -= 20
    
    # Escribir tabla
    x_offset = 30
    y_offset -= 40
    row_height = 20
    col_width = 150
    
    c.setFont("Helvetica-Bold", 12)
    for col_num, column_name in enumerate(df.columns):
        c.drawString(x_offset + col_num * col_width, y_offset, column_name)
    y_offset -= row_height
    
    c.setFont("Helvetica", 10)
    for _, row in df.iterrows():
        for col_num, value in enumerate(row):
            c.drawString(x_offset + col_num * col_width, y_offset, str(value))
        y_offset -= row_height

    c.save()
    return pdf_file

# Clase de pruebas
test_excel_file = "test_asistentes.xlsx"
test_pdf_file = "test_asistentes.pdf"

class TestPDFGenerator(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        wb = Workbook()
        ws = wb.active
        encabezado = ["Asignatura", "Fecha", "Semana Docencia", "Tema"]
        valores = ["Matemáticas", "2025-03-02", "Semana 5", "Álgebra"]
        for i, valor in enumerate(valores, start=1):
            ws[f"A{i}"] = valor
        
        datos = [["Sofía", "Aguilar Márquez", "", "Miguel", "Martín Ruiz"],
                 ["Francisco", "Blanco Medina", "", "María", "Martínez López"]]
        for i, fila in enumerate(datos, start=6):
            for j, valor in enumerate(fila, start=1):
                ws.cell(row=i, column=j, value=valor)
        
        wb.save(test_excel_file)

    @classmethod
    def tearDownClass(cls):
        os.remove(test_excel_file)
        if os.path.exists(test_pdf_file):
            os.remove(test_pdf_file)

    def test_extraer_encabezados(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Matemáticas"
        ws["A2"] = "2025-03-02"
        ws["A3"] = "Semana 5"
        ws["A4"] = "Álgebra"
        encabezados = extraer_encabezados(ws)
        self.assertEqual(encabezados["asignatura"], "Matemáticas")
        self.assertEqual(encabezados["fecha"], "2025-03-02")
        self.assertEqual(encabezados["semana_docencia"], "Semana 5")
        self.assertEqual(encabezados["tema"], "Álgebra")
    
    def test_procesar_datos(self):
        df = procesar_datos(test_excel_file)
        print("Filas cargadas:", df.shape[0])  # Depuración
        self.assertEqual(df.shape, (2, 4))  # Asegura que se carguen 2 filas y 4 columnas
    
    def test_generar_pdf(self):
        df = procesar_datos(test_excel_file)
        wb = Workbook()
        ws = wb.active
        encabezados = extraer_encabezados(ws)
        output_pdf = generar_pdf(df, encabezados, test_pdf_file)
        self.assertTrue(os.path.exists(output_pdf))

if __name__ == "__main__":
    unittest.main()
